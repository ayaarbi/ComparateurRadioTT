from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import os
import sys
import subprocess


def comparaisonAzimut(fichier_input,fichier_output,progres=None):
    """
    Compare les valeurs des colonnes M, AD et AU d'un fichier Excel et écrit le résultat dans une nouvelle colonne.
    Args:
        fichier_input (str): Chemin du fichier Excel d'entrée.
        fichier_output (str): Chemin du fichier Excel de sortie.
    """
    wb = load_workbook(fichier_input) #charger le fichier Excel d'entrée
    #On va utiliser la première feuille du fichier Excel
    ws = wb.active #Puisqu'on a une unique feuille

    #On va définir les colonnes à comparer
    col_M = column_index_from_string("M")   #corespondant au "azimut de rayonnement 2G" 
    col_AD = column_index_from_string("AD") #corespondant au "azimut de rayonnement 3G"
    col_AU = column_index_from_string("AU") #corespondant au "azimut de rayonnement 4G"

    #La colonne résultat sera ajoutée à la fin du fichier de sortie
    col_resultat = ws.max_column + 1
    ws.cell(row=1, column=col_resultat, value="Résultat de la comparaison") #ajout de la colonne
    ws.merge_cells(start_row=1, start_column=col_resultat, end_row=2, end_column=col_resultat)
    ws.cell(row=1, column=col_resultat).value = "Résultat"

    # Remplissage avec des couleur pour plus de visibilité
    vert_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") #la couleur verte pour les cellules identiques
    rouge_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") #la couleur rouge pour les cellules différentes
    
    total_rows = ws.max_row - 2  # sans compter l'entête qui est écrit sur deux lignes
    for i in range(3, ws.max_row + 1): #on commence à partir de la 3eme ligne car les 2 premiers represente l'entéte
        #Extraction des valeurs des colonnes à comparer
        azimut_2G = str(ws.cell(row=i, column=col_M).value)
        azimut_3G = str(ws.cell(row=i, column=col_AD).value)
        azimut_4G = str(ws.cell(row=i, column=col_AU).value)

        result_cell = ws.cell(row=i, column=col_resultat) #la cellule où le résultat sera écrit


        azimut_2G_list=azimut_2G.split("/") if azimut_2G else [] #On transforme les azimuts en liste pour faciliter ensuite la comparaison
        azimut_3G_list=azimut_3G.split("/") if azimut_3G else []
        azimut_4G_list=azimut_4G.split("/") if azimut_4G else []


        if (azimut_2G==azimut_3G) and ((azimut_4G is None or str(azimut_4G).strip() == "")):  #si les azimuts 2G et 3G sont égaux
            result_cell.value = "Identique"
            result_cell.fill = vert_fill
        elif azimut_2G == azimut_3G == azimut_4G: #si les trois azimuts sont égaux
            #On écrit "Identique" dans la cellule résultat et on applique le remplissage vert
            result_cell.value = "Identique"
            result_cell.fill = vert_fill
        elif azimut_2G=="Micro 2G" or azimut_2G=="Micro" or azimut_2G=="Micro " or azimut_2G=="Micro 2G/3G" or azimut_2G=="sol indoor":        # Si l'azimut 2G est "Micro 2G" 
            result_cell.value = "Identique"
            result_cell.fill = vert_fill
        elif azimut_3G=="Micro 3G" or azimut_3G=="Micro" or azimut_3G=="Micro " or azimut_3G=="Micro 2G/3G" or azimut_3G=="MICRO 3G/4G" or azimut_3G=="Micro 3G/4G" or azimut_3G=="sol indoor":  
            result_cell.value = "Identique"
            result_cell.fill = vert_fill
        elif azimut_4G=="Micro 4G":
            result_cell.value = "Identique"
            result_cell.fill = vert_fill
        
        else: 
            for j in azimut_4G_list: #si les azimutes 4G forment une combinaison des azimuts 2G et 3G
                if j in azimut_2G_list or j in azimut_3G_list:
                    result_cell.value = "Identique"
                    result_cell.fill = vert_fill
                    break
                else: #sinon on écrit "Différents" et on applique le remplissage rouge
                    result_cell.value = "Différents"  
                    result_cell.fill = rouge_fill

        if progres: #Si la fonction de progression est fournie
            #Calcul de la progression en pourcentage
            progress = int((i / total_rows) * 100)
            progres(progress)
    wb.save(fichier_output) #Enregistrer le fichier de sortie

    try:
        if sys.platform.startswith('win'): #Si l'os est windows
            os.startfile(fichier_output) #ouvrir le fichier de sortie avec Excel pour Windows
        elif sys.platform.startswith('darwin'): #Si l'os est macOS
            subprocess.call(['open', fichier_output]) 
        else: #Si l'os est l'un de distribution Linux: ubuntu, federa, debian, etc
            subprocess.call(['xdg-open', fichier_output])
    except Exception as e:
        print("Erreur d'ouverture automatique de fichier:", e) # Afficher l'erreur si l'ouverture automatique échoue


def comparaisonCoordonee(fichier_input, fichier_output, progres=None):
    """
    Compare les valeurs des colonnes E et F d'un fichier Excel et écrit le résultat dans une nouvelle colonne.
    Args:
        fichier_input (str): Chemin du fichier Excel d'entrée.
        fichier_output (str): Chemin du fichier Excel de sortie.
    """
    wb = load_workbook(fichier_input)
    ws = wb.active

    col_E = column_index_from_string("E")  # Longitude
    col_F = column_index_from_string("F")  # Latitude
    # Couleurs
    vert_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rouge_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True) #pour aligner le texte au centre

    #La colonne résultat sera ajoutée à la fin du fichier de sortie
    col_resultat = ws.max_column + 1
    ws.cell(row=1, column=col_resultat, value="Résultat de la comparaison") #ajout de la colonne
    ws.merge_cells(start_row=1, start_column=col_resultat, end_row=2, end_column=col_resultat)
    ws.cell(row=1, column=col_resultat).value = "Résultat"
    ws.cell(row=1, column=col_resultat).alignment = align_center

    

    #collecter toutes les coordonnées et les lignes associées
    coord_dict = {} 
    total_rows = ws.max_row - 2  # sans compter l'entête qui est écrit sur deux lignes

    for i in range(3, ws.max_row + 1):
        longitude = str(ws.cell(row=i, column=col_E).value).strip()
        latitude = str(ws.cell(row=i, column=col_F).value).strip()
        if longitude.lower() == "none" or latitude.lower() == "none": # Vérification des coordonnées manquantes
            cell = ws.cell(row=i, column=col_resultat)
            cell.value = "Coordonnées manquantes"
            cell.fill = rouge_fill
            cell.alignment = align_center
            continue
        coord = (longitude, latitude)  # Tuple de coordonnées
        coord_dict.setdefault(coord, []).append(i) # Ajouter la ligne à la liste des lignes pour cette coordonnée

    already_done = set()  # lignes déjà traitées
    for coord, lignes in coord_dict.items():
        if len(lignes) == 1: # Si la coordonnée est unique
            i = lignes[0]
            cell = ws.cell(row=i, column=col_resultat)
            cell.value = "Unique"
            cell.fill = vert_fill
            cell.alignment = align_center
        else: # Si la coordonnée est partagée par plusieurs lignes
            # Même message pour toutes les lignes
            lignes_str = ", ".join(map(str, lignes))
            message = f"Identique : lignes {lignes_str}"
            for i in lignes:
                cell = ws.cell(row=i, column=col_resultat)
                cell.value = message
                cell.fill = rouge_fill
                cell.alignment = align_center

        if progres:
            progress = int((len(already_done) / total_rows) * 100)
            progres(progress)
        already_done.update(lignes)

    
    # Ajuster automatiquement la largeur de la colonne résultat
    lettre_col = get_column_letter(col_resultat)
    max_length = 0

    for row in range(3, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_resultat).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))

    ws.column_dimensions[lettre_col].width = max_length + 2  # +2 pour un petit espace
    wb.save(fichier_output)

    try:
        if sys.platform.startswith('win'):
            os.startfile(fichier_output)
        elif sys.platform.startswith('darwin'):
            subprocess.call(['open', fichier_output])
        else:
            subprocess.call(['xdg-open', fichier_output])
    except Exception as e:
        print("Erreur d'ouverture automatique de fichier:", e)

